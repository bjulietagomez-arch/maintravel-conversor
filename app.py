from flask import Flask, request, send_file, jsonify, render_template
import io, re, json, base64, os
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024  # 5MB max

# Load images once at startup — write to fixed paths for openpyxl
BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
HEADER_PATH = '/tmp/mt_header.png'
FOOTER_PATH = '/tmp/mt_footer.png'
with open(os.path.join(BASE_DIR, 'static', 'header.png'), 'rb') as f:
    _hdata = f.read()
with open(os.path.join(BASE_DIR, 'static', 'footer.png'), 'rb') as f:
    _fdata = f.read()
with open(HEADER_PATH, 'wb') as f: f.write(_hdata)
with open(FOOTER_PATH, 'wb') as f: f.write(_fdata)

# ── Parser ─────────────────────────────────────────────────────────────────────
def parse_vstour(content):
    code, name = '', ''
    m = re.search(r'CLIENTE:\s*(\d+)\s+([^\r\n\\]+)', content)
    if m:
        code = m.group(1).strip()
        name = re.sub(r'\s+', ' ', m.group(2)).strip()

    rows = []
    pattern = r'\\tab\s*(\d{2}/\d{2}/\d{2,4})\s*\\tab\s*((?:FAC|REC|N/C)\s+[A-Z]\s+\d+\s+\d+)\s*\\tab\s*([\d,.()]+)\s*\\tab\s*[\d,.()]*\s*\\tab\s*([^\\]+?)\\tab\s*([^\\]+?)\\tab\s*(\d+)'
    for m in re.finditer(pattern, content):
        d, mo, y = m.group(1).split('/')
        if len(y) == 2: y = '20' + y
        fecha = datetime(int(y), int(mo), int(d))
        neg   = '(' in m.group(3)
        monto = float(m.group(3).replace('(','').replace(')','').replace(',','')) * (-1 if neg else 1)
        rows.append({
            'fecha': fecha, 'comp': m.group(2).strip(), 'monto': monto,
            'desc': m.group(4).strip(), 'pax': m.group(5).strip(), 'file': m.group(6).strip()
        })
    return {'code': code, 'name': name, 'rows': rows}

# ── Semáforo ───────────────────────────────────────────────────────────────────
def vto_color(vto: datetime):
    now   = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    day   = now.weekday()  # 0=Mon
    mon   = now - timedelta(days=day)
    sun   = mon + timedelta(days=6)
    if vto < now:  return {'bg': 'FFCCCC', 'fg': '7A0000'}
    if vto <= sun: return {'bg': 'FFF3CC', 'fg': '7A5000'}
    return             {'bg': 'CCFFCC', 'fg': '1A5C1A'}

# ── Excel builder ──────────────────────────────────────────────────────────────
def build_excel(data, credito, dias):
    NAVY, BLUE = '0D3C6C', '0F4277'
    WHITE, BLACK = 'FFFFFF', '000000'

    def fill(c):  return PatternFill('solid', fgColor=c)
    def fnt(name='Tahoma', bold=False, color=BLACK, size=10):
        return Font(name=name, bold=bold, color=color, size=size)
    def al(h='left', v='center'): return Alignment(horizontal=h, vertical=v)

    rows  = data['rows']
    n     = len(rows)
    total = sum(r['monto'] for r in rows)
    disp  = credito - total

    wb = Workbook()
    ws = wb.active
    ws.title = 'CTA CTE'

    DR = 4        # data start row (1-based)
    TR = DR+n+1   # total row
    CR = TR+1     # credito disponible
    FR = CR+2     # footer

    # ── Row 1: header image ──
    ws.row_dimensions[1].height = 60
    ws.merge_cells('A1:I1')
    ws['A1'].fill = fill(NAVY)
    try:
        img = XLImage(HEADER_PATH)
        img.anchor = 'A1'; img.width = 1000; img.height = 60
        ws.add_image(img)
    except: pass

    # ── Row 2: credito bar ──
    ws.row_dimensions[2].height = 20
    ws.merge_cells('A2:I2')
    ws['A2'] = f'CRÉDITO: ${credito:,.2f}         CRÉDITO DISPONIBLE: ${disp:,.2f}'
    ws['A2'].font      = fnt('Tahoma', bold=True, color=WHITE, size=10)
    ws['A2'].fill      = fill(NAVY)
    ws['A2'].alignment = al('right')

    # ── Row 3: cliente ──
    ws.row_dimensions[3].height = 18
    ws['B3'] = f"CLIENTE: {data['code']} {data['name']}"
    ws['B3'].font = fnt('Tahoma', size=11); ws['B3'].alignment = al('left')

    # ── Row 4: column headers ──
    ws.row_dimensions[4].height = 18
    for col, hdr in [('B','FECHA'),('C','COMPROBANTE'),('D','$'),('E','DESCRIPCIÓN'),('F','PAX'),('G','FILE'),('H','VENCIMIENTO')]:
        ws[f'{col}4'].value     = hdr
        ws[f'{col}4'].font      = fnt('Tahoma', bold=True, color=WHITE, size=9)
        ws[f'{col}4'].fill      = fill(BLUE)
        ws[f'{col}4'].alignment = al('center')

    # ── Data rows ──
    for i, row in enumerate(rows):
        r   = DR + i
        vto = row['fecha'] + timedelta(days=dias)
        col = vto_color(vto)
        ws.row_dimensions[r].height = 15

        ws[f'B{r}'] = row['fecha'];  ws[f'B{r}'].number_format = 'DD/MM/YY'; ws[f'B{r}'].alignment = al('center')
        ws[f'C{r}'] = row['comp'];   ws[f'C{r}'].alignment = al('left')
        ws[f'D{r}'] = row['monto'];  ws[f'D{r}'].number_format = '#,##0.00'; ws[f'D{r}'].alignment = al('right')
        ws[f'E{r}'] = row['desc'];   ws[f'E{r}'].alignment = al('left')
        ws[f'F{r}'] = row['pax'];    ws[f'F{r}'].alignment = al('left')
        ws[f'G{r}'] = int(row['file']) if row['file'].isdigit() else row['file']
        ws[f'G{r}'].alignment = al('center')
        ws[f'H{r}'] = vto;           ws[f'H{r}'].number_format = 'DD/MM/YY'; ws[f'H{r}'].alignment = al('center')
        ws[f'H{r}'].font = fnt('Arial Narrow', bold=True, color=col['fg'], size=9)
        ws[f'H{r}'].fill = fill(col['bg'])
        for c in ['B','C','D','E','F','G']:
            ws[f'{c}{r}'].font = fnt('Arial Narrow', size=9)

    # ── Total ──
    ws.row_dimensions[TR].height = 16
    ws[f'C{TR}'] = 'Total clientes…............'
    ws[f'C{TR}'].font = fnt('Tahoma', bold=True, size=10); ws[f'C{TR}'].alignment = al('right')
    ws[f'D{TR}'] = total
    ws[f'D{TR}'].number_format = '#,##0.00'; ws[f'D{TR}'].font = fnt('Tahoma', bold=True, size=10); ws[f'D{TR}'].alignment = al('right')

    # ── Crédito disponible ──
    ws.row_dimensions[CR].height = 16
    ws[f'C{CR}'] = 'Crédito Disponible'
    ws[f'C{CR}'].font = fnt('Tahoma', bold=True, color=NAVY, size=10); ws[f'C{CR}'].alignment = al('right')
    ws[f'D{CR}'] = disp
    ws[f'D{CR}'].number_format = '#,##0.00'
    ws[f'D{CR}'].font = fnt('Tahoma', bold=True, color=('AE0000' if disp < 0 else '1A7A4A'), size=10)
    ws[f'D{CR}'].alignment = al('right')

    # ── Footer image ──
    ws.row_dimensions[FR].height = 50
    ws.merge_cells(f'A{FR}:I{FR}')
    ws[f'A{FR}'].fill = fill(NAVY)
    try:
        fimg = XLImage(FOOTER_PATH)
        fimg.anchor = f'A{FR}'; fimg.width = 1000; fimg.height = 50
        ws.add_image(fimg)
    except: pass

    # Column widths
    for col, w in [('A',2),('B',11),('C',24),('D',16),('E',30),('F',24),('G',8),('H',13),('I',2)]:
        ws.column_dimensions[col].width = w

    # ── Sheet 2: FORMAS DE PAGO ──
    ws2 = wb.create_sheet('FORMAS DE PAGO')
    ws2.row_dimensions[1].height = 60
    ws2.merge_cells('A1:I1')
    ws2['A1'].fill = fill(NAVY)
    try:
        himg2 = XLImage(HEADER_PATH); himg2.anchor='A1'; himg2.width=1000; himg2.height=60
        ws2.add_image(himg2)
    except: pass

    ws2.row_dimensions[2].height = 20
    ws2.merge_cells('A2:I2')
    ws2['A2'] = f"CLIENTE: {data['code']} {data['name']}"
    ws2['A2'].font = fnt('Tahoma', bold=True, color=WHITE, size=10)
    ws2['A2'].fill = fill(NAVY); ws2['A2'].alignment = al('left')

    fp_data = [
        (4,'D','PESOS ARGENTINOS',True),(4,'F','DÓLARES USD',True),
        (5,'D','TITULAR: MAINTRAVEL International SRL',False),(5,'F','TITULAR: MAINTRAVEL International SRL',False),
        (6,'D','CUIT: 30-70983411-4',False),(6,'F','CUIT: 30-70983411-4',False),
        (7,'D','CUENTA CORRIENTE EN PESOS: 1784/2 SUC 204',False),(7,'F','CUENTA CORRIENTE EN DÓLARES: 2908/5 suc. 204',False),
        (8,'D','CBU: 07202041-20000000178422',False),(8,'F','CBU: 07202041-21000000290855',False),
        (9,'D','ALIAS: SANTANDERMTPESOS',False),(9,'F','ALIAS: SANTANDERMTDOLARES',False),
        (11,'D','TARJETAS DE CRÉDITO',True),
        (12,'D','Operaciones con tarjetas de crédito/débito: gasto administrativo del 4%.',False),
        (13,'D','AHORA 12 Y AHORA 18 — Exclusivo para productos seleccionados.',False),
    ]
    for rn, co, val, bold in fp_data:
        ws2[f'{co}{rn}'] = val
        ws2[f'{co}{rn}'].font = fnt('Tahoma', bold=bold, size=10)

    ws2.row_dimensions[16].height = 50
    ws2.merge_cells('A16:I16')
    ws2['A16'].fill = fill(NAVY)
    try:
        fimg2 = XLImage(FOOTER_PATH); fimg2.anchor='A16'; fimg2.width=1000; fimg2.height=50
        ws2.add_image(fimg2)
    except: pass

    for col, w in [('A',2),('B',11),('C',24),('D',16),('E',30),('F',24),('G',8),('H',13),('I',2)]:
        ws2.column_dimensions[col].width = w


    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── Routes ─────────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/convert', methods=['POST'])
def convert():
    try:
        file    = request.files.get('file')
        credito = float(request.form.get('credito', 10000000))
        dias    = int(request.form.get('dias', 15))

        if not file:
            return jsonify({'error': 'No se recibió archivo'}), 400

        content = file.read().decode('latin-1', errors='ignore')
        data    = parse_vstour(content)

        if not data['rows']:
            return jsonify({'error': 'No se encontraron transacciones en el archivo'}), 400

        buf  = build_excel(data, credito, dias)
        name = f"Estado_de_Cuenta_{data['name'].replace(' ','_')[:20]}_{datetime.now().strftime('%d%m%Y')}.xlsx"

        return send_file(buf, as_attachment=True, download_name=name,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/preview', methods=['POST'])
def preview():
    try:
        file = request.files.get('file')
        if not file:
            return jsonify({'error': 'No se recibió archivo'}), 400
        content = file.read().decode('latin-1', errors='ignore')
        data    = parse_vstour(content)
        if not data['rows']:
            return jsonify({'error': 'No se encontraron transacciones'}), 400
        # Serialize dates as strings
        rows_out = []
        for r in data['rows']:
            rows_out.append({**r, 'fecha': r['fecha'].strftime('%d/%m/%y')})
        return jsonify({'code': data['code'], 'name': data['name'], 'rows': rows_out})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
